#!/usr/bin/env python
# coding:utf-8

#!/usr/bin/env python
# coding:utf-8

import pyttsx3
import openpyxl
import random
import tkinter as tk
import tkinter.messagebox
import os

path = os.getcwd()
print(path)
def voice(a): #语音读取被抽到的学生的姓名
    ## pytttsx初始化
    engine = pyttsx3.init()
    ## a代表着需要转成音频的文字
    engine.say(a)
    # 注意，没有本句话是没有声音的
    engine.runAndWait()

### 创建一个list来保存已经被抽到名字的同学
ed_name= []

m = tk.Tk()  # 创建窗口对象
m.title("点兵点将 v1.0 ")
m.geometry("600x400")

# 创建抽取到的同学姓名展示部分
labelx=tk.Label(m,text="点击下方",fg="red",font=("宋体",80),width=9,height=2).grid(row=0)

name_data = [] # 创建存储所有学生姓名的列表

def execl(filename, sheetname):
    wb = openpyxl.load_workbook(filename)  # 打开Excel
    sheet = wb[sheetname]  # 定位表单
    # 创建一个空列表
    column = 1
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row, column).value == "":
            continue
        name_data.append(sheet.cell(row, column).value)  # 将第一列的每一行数值遍历添加到name_data列表中；
    return name_data
execl(r"name\name.xlsx","Sheet1")  # 调用一下execl函数，在没有开始点名之前先将姓名添加到name_data列表之中


def callname():
    while 1:
        a = random.randint(0,len(name_data)-1)## 产生随机数，作为list的下标index

        b = "下面请"+name_data[a]+"同学来回答问题"## 拼接成字符串

        ## 判断ed_name元素个数是不是与name_data列表中相同，如果相同，那么代表着所有的同学名字都被抽点过了
        if len(ed_name) == len(name_data):
            tk.messagebox.askokcancel("提示", '所有的学生都已经抽取了一遍，请重启程序以重新开始抽取')
            break

        ## 判断被抽点到的同学是不是已经被抽点过，如果被抽点过，则跳过该同学
        if name_data[a] in ed_name:
            continue

        ed_name.append(name_data[a])    # 将被抽点到的学生添加到ed_name例表中，标记已经被抽点

        labelx = tk.Label(m,text=name_data[a],fg="red",font=("微软雅黑", 80),width=9,height=2).grid(row=0) #展示被抽取到的学生的姓名

        voice(b) ## 语音读取

        break

tk.Button(m,text="开始点名",width=15,command=callname).grid(row=1,padx=10,pady=10,sticky='s')  # s南边也就是下边


m.mainloop()  # 进入消息循环

